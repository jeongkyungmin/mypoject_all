from openpyxl import load_workbook

# data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook("D:\Dsparta\mypoject_all/fruit.xlsx", data_only=True)
# 시트 이름으로 불러오기
load_ws = load_wb['sheet1']

# 셀 주소로 값 출력
print(load_ws['A1'].value)

# 셀 좌표로 값 출력
print(load_ws.cell(1, 2).value)

from openpyxl import Workbook

write_wb = Workbook()

# 이름이 있는 시트를 생성
# write_ws = write_wb.create_sheet('생성시트')

# Sheet1에다 입력
write_ws = write_wb.active
write_ws['A1'] = '숫자'

# 행 단위로 추가
write_ws.append([1, 2, 3])

# 셀 단위로 추가
write_ws.cell(5, 5, '5행5열')
write_wb.save('D:\Dsparta\mypoject_all/number.xlsx')